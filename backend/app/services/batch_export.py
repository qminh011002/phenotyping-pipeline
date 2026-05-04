"""Build a ZIP archive of a batch's reviewed images + an .xlsx summary.

The archive is streamed back to the client so we never hold the whole zip in
memory at once. Per-image inclusion is decided by the caller (the download
dialog sends the subset of image IDs the user ticked).

Image source: the raw upload with the operator's reviewed boxes drawn on top
— matching what the user saw in the ResultViewer when they clicked Finish.
Edited boxes win when present; otherwise the model's annotations are drawn.
We deliberately do NOT export the script-generated overlay PNG (which carries
the inference-time Configuration/Result text boards), because that is not
what the operator reviewed.

Summary columns reflect the *edited* count/average confidence when the
operator has saved edits, otherwise the model's output.
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
import tempfile
import zipfile
from collections.abc import AsyncIterator, Iterable
from pathlib import Path
from uuid import UUID

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.analysis import AnalysisBatch, AnalysisImage

logger = logging.getLogger(__name__)

# Excel styling — a calm, "corporate report" look: dark header with white
# bold text, zebra-striped body, thin borders, auto-sized columns.
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # slate-800
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="111827")
_META_FONT = Font(name="Calibri", size=10, color="4B5563")  # slate-600
_ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")  # slate-50
_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
_SUMMARY_HEADERS = [
    "#",
    "Filename",
    "Count",
    "Avg confidence",
    "Elapsed (s)",
    "Edited",
]
# Columns are typed so numeric formats render correctly in Excel.
_NUMERIC_FORMATS = {
    3: "#,##0",  # Count
    4: "0.0%",  # Avg confidence
    5: "0.00",  # Elapsed seconds
}


def _slugify(value: str) -> str:
    """Filesystem-safe filename slug. Keeps letters/digits/.-_ and collapses
    everything else to '_'."""
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._")
    return clean or "batch"


def _effective_stats(image: AnalysisImage) -> tuple[int, float | None, bool]:
    """Return (count, avg_confidence, edited).

    When `edited_annotations` is populated, the counts reflect the operator's
    edits; otherwise we fall back to the stored model totals. User-drawn boxes
    may not carry a meaningful confidence — they're excluded from the average
    but included in the count, matching how the ResultViewer's StatBoard
    behaves.
    """
    edited = image.edited_annotations
    if isinstance(edited, list) and len(edited) > 0:
        count = len(edited)
        confidences: list[float] = []
        for box in edited:
            if not isinstance(box, dict):
                continue
            origin = box.get("origin", "model")
            if origin == "user":
                continue
            conf = box.get("confidence")
            if isinstance(conf, (int, float)):
                confidences.append(float(conf))
        avg = sum(confidences) / len(confidences) if confidences else None
        return count, avg, True

    # No edits — fall back to the stored aggregates.
    return (image.count or 0, image.avg_confidence, False)


def _style_summary_sheet(
    ws,
    batch: AnalysisBatch,
    rows: list[tuple[int, str, int, float | None, float | None, bool]],
) -> None:
    """Write a formatted summary sheet: batch header → meta → table."""
    # Title row
    ws["A1"] = batch.name or "Untitled batch"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(_SUMMARY_HEADERS)
    )
    ws.row_dimensions[1].height = 24

    # Meta rows (organism / device / mode / totals)
    meta_pairs = [
        ("Organism", (batch.organism_type or "").capitalize()),
        ("Device", (batch.device or "").upper()),
        ("Mode", (batch.mode or "").capitalize()),
        ("Images exported", len(rows)),
        ("Total count", batch.total_count if batch.total_count is not None else "—"),
        (
            "Avg confidence",
            (
                f"{batch.avg_confidence * 100:.1f}%"
                if batch.avg_confidence is not None
                else "—"
            ),
        ),
    ]
    for i, (label, value) in enumerate(meta_pairs, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(
            name="Calibri", size=10, bold=True, color="4B5563"
        )
        ws.cell(row=i, column=2, value=value).font = _META_FONT

    # Blank spacer row, then the table header
    header_row = 2 + len(meta_pairs) + 1
    for col, label in enumerate(_SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[header_row].height = 22

    # Body rows
    data_start = header_row + 1
    for i, (idx, filename, count, avg_conf, elapsed, edited) in enumerate(rows):
        r = data_start + i
        zebra = i % 2 == 1

        values = [
            idx,
            filename,
            count,
            avg_conf,  # may be None
            elapsed,  # may be None
            "Yes" if edited else "No",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = _BORDER
            if zebra:
                cell.fill = _ZEBRA_FILL
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col == 2:
                cell.alignment = Alignment(horizontal="left")
                cell.font = Font(name="Consolas", size=10)
            else:
                cell.alignment = Alignment(horizontal="right")
            fmt = _NUMERIC_FORMATS.get(col)
            if fmt is not None:
                cell.number_format = fmt

    # Column widths — big enough for typical filenames, compact elsewhere.
    widths = {1: 5, 2: 42, 3: 10, 4: 16, 5: 14, 6: 10}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze the title/meta + header so scrolling keeps them visible.
    ws.freeze_panes = ws.cell(row=data_start, column=1)


def _build_xlsx(batch: AnalysisBatch, images: Iterable[AnalysisImage]) -> bytes:
    """Render the styled summary workbook to an in-memory bytes buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    rows: list[tuple[int, str, int, float | None, float | None, bool]] = []
    for idx, img in enumerate(images, start=1):
        count, avg_conf, edited = _effective_stats(img)
        rows.append(
            (idx, img.original_filename, count, avg_conf, img.elapsed_secs, edited)
        )

    _style_summary_sheet(ws, batch, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolve_overlay_path(overlay_path: str, storage_dir: Path) -> Path:
    p = Path(overlay_path)
    if not p.is_absolute():
        p = storage_dir / p
    return p


def _find_raw_path(overlay_path: Path) -> Path | None:
    """Locate the raw upload that sits next to the overlay PNG.

    The inference service writes ``{filename}_raw{original_suffix}`` alongside
    ``{filename}_overlay.png``; the suffix preserves the original upload's
    extension so we glob for it.
    """
    raw_stem = overlay_path.name.replace("_overlay.png", "_raw")
    candidates = sorted(overlay_path.parent.glob(f"{raw_stem}.*"))
    return candidates[0] if candidates else None


def _boxes_for_render(image: AnalysisImage) -> list[dict]:
    """Pick which annotation list to draw — edits win, fall back to model."""
    edited = image.edited_annotations
    if isinstance(edited, list) and len(edited) > 0:
        return [b for b in edited if isinstance(b, dict)]
    annotations = image.annotations
    if isinstance(annotations, list):
        return [b for b in annotations if isinstance(b, dict)]
    return []


def _coerce_bbox(box: dict) -> tuple[int, int, int, int] | None:
    """Pull (x1, y1, x2, y2) out of an annotation dict, in pixel coordinates."""
    raw = box.get("bbox")
    if not isinstance(raw, (list, tuple)) or len(raw) != 4:
        return None
    try:
        x1, y1, x2, y2 = (int(round(float(v))) for v in raw)
    except (TypeError, ValueError):
        return None
    if x2 < x1:
        x1, x2 = x2, x1
    if y2 < y1:
        y1, y2 = y2, y1
    return x1, y1, x2, y2


def _render_reviewed_png(raw_path: Path, boxes: list[dict]) -> bytes | None:
    """Draw `boxes` on the raw image and return PNG-encoded bytes.

    Box style mirrors the inference overlay: 1-px green rectangle with the
    confidence printed just above the top-left corner. User-drawn boxes that
    don't carry a numeric confidence are drawn without a label.

    Returns None when the raw image can't be decoded.
    """
    data = np.fromfile(str(raw_path), dtype=np.uint8)
    if data.size == 0:
        return None
    image = cv2.imdecode(data, cv2.IMREAD_COLOR)
    if image is None:
        return None

    color = (0, 255, 0)
    for box in boxes:
        coords = _coerce_bbox(box)
        if coords is None:
            continue
        x1, y1, x2, y2 = coords
        cv2.rectangle(image, (x1, y1), (x2, y2), color, 1)
        conf = box.get("confidence")
        if isinstance(conf, (int, float)) and box.get("origin") != "user":
            cv2.putText(
                image,
                f"{float(conf):.2f}",
                (x1, max(y1 - 5, 0)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.4,
                color,
                1,
            )

    ok, buf = cv2.imencode(".png", image, [cv2.IMWRITE_PNG_COMPRESSION, 1])
    if not ok:
        return None
    return buf.tobytes()


async def build_batch_archive(
    batch_id: UUID,
    image_ids: list[UUID] | None,
    db: AsyncSession,
    storage_dir: Path,
    user_id: UUID,
) -> tuple[str, "tempfile.SpooledTemporaryFile"] | None:
    """Build the ZIP archive for a batch.

    Returns (filename, spooled_file) — the file is positioned at offset 0 and
    the caller is responsible for closing it. Returns None when the batch
    doesn't exist. Raises ValueError if ``image_ids`` is supplied but none
    belong to the batch.

    The archive spools to RAM up to 100 MB, then spills to disk, so peak RSS
    is bounded regardless of how many overlays a batch holds.
    """
    batch_stmt = (
        select(AnalysisBatch)
        .where(AnalysisBatch.id == batch_id)
        .where(AnalysisBatch.user_id == user_id)
    )
    batch = (await db.execute(batch_stmt)).scalar_one_or_none()
    if batch is None:
        return None

    img_stmt = (
        select(AnalysisImage)
        .where(AnalysisImage.batch_id == batch_id)
        .where(AnalysisImage.status == "completed")
        .order_by(AnalysisImage.created_at)
    )
    if image_ids:
        img_stmt = img_stmt.where(AnalysisImage.id.in_(image_ids))

    images: list[AnalysisImage] = list((await db.execute(img_stmt)).scalars().all())
    if not images:
        raise ValueError(
            "No completed images match the requested selection for this batch."
        )

    def _build(out: tempfile.SpooledTemporaryFile) -> None:
        with zipfile.ZipFile(out, mode="w", compression=zipfile.ZIP_STORED) as zf:
            stem_counts: dict[str, int] = {}
            for img in images:
                if not img.overlay_path:
                    continue
                overlay_src = _resolve_overlay_path(img.overlay_path, storage_dir)
                raw_src = _find_raw_path(overlay_src)

                rendered: bytes | None = None
                if raw_src is not None and raw_src.exists():
                    rendered = _render_reviewed_png(raw_src, _boxes_for_render(img))

                if rendered is None:
                    # Fall back to the script overlay if we can't render the
                    # reviewed view (raw missing, decode failure, …) so the
                    # export still produces something useful.
                    if not overlay_src.exists():
                        logger.warning(
                            "Skipping image during export — no raw or overlay on disk",
                            extra={
                                "context": {
                                    "batch_id": str(batch_id),
                                    "image_id": str(img.id),
                                    "overlay_path": str(overlay_src),
                                }
                            },
                        )
                        continue
                    logger.warning(
                        "Falling back to script overlay during export",
                        extra={
                            "context": {
                                "batch_id": str(batch_id),
                                "image_id": str(img.id),
                                "raw_path": str(raw_src) if raw_src else None,
                            }
                        },
                    )

                stem = Path(img.original_filename).stem
                count = stem_counts.get(stem, 0)
                if rendered is not None:
                    arc_ext = ".png"
                else:
                    arc_ext = overlay_src.suffix.lower() or ".png"
                candidate = (
                    f"{stem}{arc_ext}" if count == 0 else f"{stem}_{count}{arc_ext}"
                )
                stem_counts[stem] = count + 1

                if rendered is not None:
                    info = zipfile.ZipInfo(f"images/{candidate}")
                    info.compress_type = zipfile.ZIP_STORED
                    zf.writestr(info, rendered)
                else:
                    zf.write(overlay_src, arcname=f"images/{candidate}")

            xlsx_bytes = _build_xlsx(batch, images)
            xlsx_info = zipfile.ZipInfo("summary.xlsx")
            xlsx_info.compress_type = zipfile.ZIP_DEFLATED
            zf.writestr(xlsx_info, xlsx_bytes)

    # Spool to RAM up to ~100 MB; spill to disk above that. Bounds peak RSS
    # for batches with many large overlays.
    spool: tempfile.SpooledTemporaryFile = tempfile.SpooledTemporaryFile(
        max_size=100 * 1024 * 1024
    )
    await asyncio.to_thread(_build, spool)
    spool.seek(0)
    filename = f"{_slugify(batch.name or 'batch')}.zip"
    return filename, spool


async def stream_batch_archive(
    batch_id: UUID,
    image_ids: list[UUID] | None,
    db: AsyncSession,
    storage_dir: Path,
    user_id: UUID,
    chunk_size: int = 64 * 1024,
) -> tuple[str, AsyncIterator[bytes]] | None:
    """Wrapper that yields the archive bytes in chunks for StreamingResponse."""
    built = await build_batch_archive(
        batch_id=batch_id,
        image_ids=image_ids,
        db=db,
        storage_dir=storage_dir,
        user_id=user_id,
    )
    if built is None:
        return None
    filename, spool = built

    async def _iter() -> AsyncIterator[bytes]:
        try:
            while True:
                # SpooledTemporaryFile read may touch disk once spilled, so do
                # it in a worker thread to keep the loop responsive.
                chunk = await asyncio.to_thread(spool.read, chunk_size)
                if not chunk:
                    return
                yield chunk
        finally:
            spool.close()

    return filename, _iter()
